#!/usr/bin/env python3
"""analysis-to-xlsx.py — turn a twt-text-analysis analysis-report.md into an XLSX.

The report writes each block as a `## Block N -- Type` section followed by a fixed
run of labelled fields (Purpose, Original, Applicable Metrics, Overall, Finding
Type, Decision, Weaknesses, Can Fix Safely, Reason, Suggested Version, Rewrite
Validation, Confidence). This script parses those sections deterministically and
emits one row per block with the analytical columns, colour-coding the finding.

Usage:
  python analysis-to-xlsx.py --input <analysis-report.md> [--output <file.xlsx>]

Exit codes: 0 ok, 1 usage/parse error. Depends only on openpyxl.
"""
import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: python -m pip install openpyxl\n")
    sys.exit(1)

LABELS = [
    "Purpose", "Original", "Applicable Metrics", "Overall", "Finding Type",
    "Decision", "Weaknesses", "Can Fix Safely", "Reason", "Suggested Version",
    "Weakness-To-Fix Mapping", "Rewrite Validation", "Confidence",
]
LABEL_RE = re.compile(r"^(" + "|".join(re.escape(l) for l in LABELS) + r"):\s*(.*)$")
BLOCK_RE = re.compile(r"^##\s+Block\s+(\d+)\s*[—–-]\s*(.+?)\s*$")

# Output columns: (header, field-key or special)
COLUMNS = [
    ("Block", "_num"),
    ("Type", "_type"),
    ("Score", "Overall"),
    ("Finding Type", "Finding Type"),
    ("Original", "Original"),
    ("Suggested Version", "Suggested Version"),
    ("Weaknesses", "Weaknesses"),
    ("Can Fix Safely", "Can Fix Safely"),
    ("Reason", "Reason"),
    ("Rewrite Validation", "Rewrite Validation"),
    ("Confidence", "Confidence"),
]


def clean_value(label, lines):
    """Collapse a field's raw lines into a cell string, format-aware per label."""
    text = "\n".join(lines).strip()
    # Strip a fenced code block wrapper, keep the inner content.
    fence = re.match(r"^```[^\n]*\n(.*?)\n?```$", text, re.DOTALL)
    if fence:
        text = fence.group(1)
    # Inline-code wrapper around a single value.
    if text.startswith("`") and text.endswith("`") and text.count("`") == 2:
        text = text[1:-1]
    if label in ("Weaknesses", "Rewrite Validation", "Applicable Metrics"):
        # Keep as one bullet per line, marker stripped.
        rows = [re.sub(r"^\s*[-*•]\s+", "", ln).strip() for ln in text.splitlines()]
        return "\n".join(r for r in rows if r)
    if label in ("Original", "Suggested Version"):
        return text  # preserve internal newlines verbatim
    # Prose fields: soft-wrap newlines into spaces.
    return re.sub(r"\s*\n\s*", " ", text).strip()


def parse_blocks(md):
    lines = md.splitlines()
    blocks = []
    i = 0
    n = len(lines)
    while i < n:
        m = BLOCK_RE.match(lines[i])
        if not m:
            i += 1
            continue
        block = {"_num": m.group(1), "_type": m.group(2).strip()}
        i += 1
        cur_label = None
        cur_lines = []

        def flush():
            if cur_label is not None:
                block[cur_label] = clean_value(cur_label, cur_lines)

        while i < n and not BLOCK_RE.match(lines[i]):
            line = lines[i]
            lm = LABEL_RE.match(line)
            if lm:
                flush()
                cur_label = lm.group(1)
                cur_lines = [lm.group(2)] if lm.group(2).strip() else []
            elif cur_label is not None:
                cur_lines.append(line)
            i += 1
        flush()
        blocks.append(block)
    return blocks


def finding_fill(value):
    v = (value or "").strip().lower()
    if v == "problem":
        return PatternFill("solid", fgColor="F7D4D2")   # soft red
    if v == "opportunity":
        return PatternFill("solid", fgColor="FBEBC8")   # soft amber
    if v == "no issue":
        return PatternFill("solid", fgColor="D7EBD8")   # soft green
    return None


def build_workbook(blocks, subject):
    wb = Workbook()
    ws = wb.active
    ws.title = "Text Analysis"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="222741")
    thin = Side(style="thin", color="D9DCEC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(vertical="top", wrap_text=True)

    # Header row
    for c, (title, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border

    # Data rows
    for r, block in enumerate(blocks, start=2):
        for c, (_title, key) in enumerate(COLUMNS, start=1):
            val = block.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap_top
            cell.border = border
            if key == "Finding Type":
                fill = finding_fill(val)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)

    widths = {
        "Block": 7, "Type": 26, "Score": 9, "Finding Type": 14, "Original": 44,
        "Suggested Version": 40, "Weaknesses": 42, "Can Fix Safely": 12,
        "Reason": 52, "Rewrite Validation": 30, "Confidence": 11,
    }
    for c, (title, _key) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(title, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(blocks) + 1}"
    ws.sheet_properties.tabColor = "222741"
    if subject:
        ws.title = "Text Analysis"
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output")
    args = ap.parse_args()

    src = Path(args.input)
    if not src.is_file():
        sys.stderr.write(f"Input not found: {src}\n")
        sys.exit(1)
    md = src.read_text(encoding="utf-8")
    subject_m = re.search(r"^-\s+\*\*Subject:\*\*\s+(.*)$", md, re.MULTILINE)
    subject = subject_m.group(1).strip() if subject_m else ""

    blocks = parse_blocks(md)
    if not blocks:
        sys.stderr.write("No '## Block N — Type' sections found; nothing to export.\n")
        sys.exit(1)

    out = Path(args.output) if args.output else src.with_suffix(".xlsx")
    wb = build_workbook(blocks, subject)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"XLSX written: {out}")
    print(f"Rows: {len(blocks)} block(s)")


if __name__ == "__main__":
    main()
