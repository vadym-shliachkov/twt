#!/usr/bin/env python3
"""checklist-xlsx.py — deterministic content-approval workbook builder + reader.

/twt-content-approval-checklist used to hand-write openpyxl code on every run to
implement ~15 precise styling rules (merged banner rows, zebra shading restarting
per section, the green/pink ready cell, data validation, frozen header, widths) —
exactly the format-drift risk analysis-to-xlsx.py already solved for text-analysis.
This script owns the workbook MECHANICS; the model owns the judgment (content
discovery, block/field decisions) and hands it over as a JSON spec.

Usage:
  python checklist-xlsx.py build --spec <spec.json> --out <workbook.xlsx>
  python checklist-xlsx.py read  --workbook <workbook.xlsx>

Build spec shape (worksheet order is preserved — put `Shared header` first and
`Shared footer` last, per the checklist skill):
  {
    "worksheets": [
      { "name": "Shared header",
        "blocks": [
          { "name": "Primary navigation",
            "rows": [
              { "field_type": "text:logo_text", "current": "Acme",
                "recommended": "Needs approved final copy" }
            ] } ] }
    ]
  }
Row keys: field_type (required), current, recommended, approved (optional,
normally blank), ready (optional, defaults false).

Read mode emits a fenced JSON block: every field row per worksheet (banner and
spacer rows skipped silently — a blank field type marks a banner), `ready`
normalized (true/yes/1/TRUE → true), `implementable` = ready AND approved
non-blank, `family` = the field_type prefix, plus a summary and a `duplicates`
list (same sheet+block+field_type with conflicting non-blank approved values).

Exit codes: 0 ok, 1 usage/parse error. Depends only on openpyxl.
"""
import argparse
import json
import sys
from pathlib import Path

# Windows consoles default to cp1252 — force UTF-8 so approved copy with
# em-dashes/quotes survives the JSON round-trip to the caller.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: python -m pip install openpyxl\n")
    sys.exit(1)

HEADERS = [
    "Block name", "field type", "current content",
    "recommended content", "approved content", "ready to implement (true, false)",
]
NCOLS = len(HEADERS)
READY_COL = get_column_letter(NCOLS)  # F
WIDTHS = [24, 22, 46, 46, 46, 14]

# Palette (matches analysis-to-xlsx.py)
HEADER_FILL = "222741"
BANNER_FILL = "2F3561"
ZEBRA_FILL = "F4F5FA"
GREEN_FILL = "D7EBD8"
PINK_FILL = "F7D4D2"

THIN = Side(style="thin", color="D9DCEC")
THICK = Side(style="thick", color=BANNER_FILL)
FIELD_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)


def build(spec_path, out_path):
    try:
        spec = json.loads(Path(spec_path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        sys.stderr.write(f"Cannot read spec {spec_path}: {e}\n")
        sys.exit(1)
    worksheets = spec.get("worksheets") or []
    if not worksheets:
        sys.stderr.write("Spec has no worksheets.\n")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)
    summary = []

    for sheet in worksheets:
        name = str(sheet.get("name", "")).strip()
        if not name:
            sys.stderr.write("A worksheet is missing its name.\n")
            sys.exit(1)
        ws = wb.create_sheet(title=name[:31])  # Excel sheet-name limit

        # Header row
        for c, title in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(vertical="center", horizontal="left")
            cell.border = FIELD_BORDER
        ws.freeze_panes = "A2"
        for c, w in enumerate(WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

        # Data validation for the ready column (field rows join it as written)
        dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        ws.add_data_validation(dv)

        r = 2
        nrows = 0
        blocks = sheet.get("blocks") or []
        for block in blocks:
            bname = str(block.get("name", "")).strip() or "Block"

            # Section banner row: merged, bold, white on accent, thick top border.
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
            banner = ws.cell(row=r, column=1, value=bname)
            banner.font = Font(bold=True, color="FFFFFF", size=12)
            banner.alignment = Alignment(vertical="center", horizontal="left")
            for c in range(1, NCOLS + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=BANNER_FILL)
                cell.border = Border(top=THICK)
            ws.row_dimensions[r].height = 22
            r += 1

            # Field rows: zebra restarts at every banner.
            for i, row in enumerate(block.get("rows") or []):
                zebra = i % 2 == 1
                values = [
                    bname,
                    str(row.get("field_type", "")).strip(),
                    row.get("current", "") or "",
                    row.get("recommended", "") or "",
                    row.get("approved", "") or "",
                    "true" if row.get("ready") in (True, "true", "TRUE", "yes", 1) else "false",
                ]
                if not values[1]:
                    sys.stderr.write(f"Sheet '{name}' block '{bname}' row {i + 1} has no field_type.\n")
                    sys.exit(1)
                for c, val in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.alignment = WRAP_TOP
                    cell.border = FIELD_BORDER
                    if zebra:
                        cell.fill = PatternFill("solid", fgColor=ZEBRA_FILL)
                # De-emphasize the repeated block name; the banner is the label.
                ws.cell(row=r, column=1).font = Font(color="8A8FA8", size=10)
                dv.add(ws.cell(row=r, column=NCOLS))
                r += 1
                nrows += 1

            # Spacer row: pure whitespace (no fill, no border).
            r += 1

        # Ready-cell fill by literal value, via conditional formatting so it
        # re-evaluates when a reviewer edits the cell. Excel string comparison
        # is case-insensitive, and =TRUE covers boolean entry; blank banner /
        # spacer cells match neither rule.
        if r > 2:
            rng = f"{READY_COL}2:{READY_COL}{r - 1}"
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'OR(${READY_COL}2="true",${READY_COL}2=TRUE)'],
                fill=PatternFill("solid", fgColor=GREEN_FILL)))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'OR(${READY_COL}2="false",${READY_COL}2=FALSE)'],
                fill=PatternFill("solid", fgColor=PINK_FILL)))

        summary.append({"worksheet": ws.title, "blocks": len(blocks), "rows": nrows})

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Workbook written: {out}")
    print("```json")
    print(json.dumps({"workbook": str(out), "worksheets": summary}, indent=2))
    print("```")


def norm_ready(v):
    if v is True:
        return True
    return str(v or "").strip().lower() in ("true", "yes", "1")


def read(workbook_path):
    wb_path = Path(workbook_path)
    if not wb_path.is_file():
        sys.stderr.write(f"Workbook not found: {wb_path}\n")
        sys.exit(1)
    wb = load_workbook(wb_path, data_only=True)

    worksheets = []
    duplicates = []
    seen = {}  # (sheet, block, field_type) -> first non-blank approved value
    total = implementable = not_ready = blank_approved = 0

    for ws in wb.worksheets:
        rows = []
        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(row=r, column=c).value for c in range(1, NCOLS + 1)]
            block, ftype, current, recommended, approved, ready_raw = (
                (str(v).strip() if v is not None and v is not True and v is not False else v)
                for v in cells
            )
            block = block or ""
            ftype = ftype or ""
            if not ftype:
                continue  # banner or spacer row — skip silently
            ready = norm_ready(ready_raw)
            approved_s = "" if approved is None else str(approved).strip()
            impl = ready and approved_s != ""
            total += 1
            if impl:
                implementable += 1
            elif not ready:
                not_ready += 1
            else:
                blank_approved += 1
            family = ftype.split(":", 1)[0] if ":" in ftype else ftype
            key = (ws.title, block, ftype)
            if approved_s:
                if key in seen and seen[key] != approved_s:
                    duplicates.append({
                        "worksheet": ws.title, "block": block, "field_type": ftype,
                        "values": [seen[key], approved_s],
                    })
                else:
                    seen.setdefault(key, approved_s)
            rows.append({
                "row": r, "block": block, "field_type": ftype, "family": family,
                "current": current or "", "recommended": recommended or "",
                "approved": approved_s, "ready": ready, "implementable": impl,
            })
        worksheets.append({"name": ws.title, "rows": rows})

    print(f"checklist-read: {total} field rows, {implementable} implementable, "
          f"{not_ready} not ready, {blank_approved} ready-but-blank, {len(duplicates)} conflict(s)")
    print("```json")
    print(json.dumps({
        "workbook": str(wb_path),
        "summary": {"total_rows": total, "implementable": implementable,
                    "not_ready": not_ready, "ready_but_blank_approved": blank_approved},
        "duplicates": duplicates,
        "worksheets": worksheets,
    }, indent=2, ensure_ascii=False))
    print("```")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="mode", required=True)
    b = sub.add_parser("build")
    b.add_argument("--spec", required=True)
    b.add_argument("--out", required=True)
    rd = sub.add_parser("read")
    rd.add_argument("--workbook", required=True)
    args = ap.parse_args()
    if args.mode == "build":
        build(args.spec, args.out)
    else:
        read(args.workbook)


if __name__ == "__main__":
    main()
